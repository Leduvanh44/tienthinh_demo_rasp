import os
import time
import schedule
from datetime import datetime
from openpyxl import Workbook, load_workbook

def get_cpu_temperature():
    with open("/sys/class/thermal/thermal_zone0/temp", "r") as f:
        temp = int(f.read()) / 1000.0  
    return temp


def log_temperature_to_excel():

    file_name = "cpu_temperature_log.xlsx"

    today = datetime.now().strftime("%Y-%m-%d")

    time_now = datetime.now().strftime("%H:%M:%S")
    

    if os.path.exists(file_name):
        workbook = load_workbook(file_name)
    else:
        workbook = Workbook()
    
    if today not in workbook.sheetnames:
        sheet = workbook.create_sheet(title=today)
        sheet.append(["Thời gian hiện tại", "Nhiệt độ (°C)"])
    else:
        sheet = workbook[today]
    

    temperature = get_cpu_temperature()
    sheet.append([time_now, temperature])
    print(f"[{time_now}] Nhiệt độ CPU: {temperature}°C - Đã lưu vào sheet {today}")
    

    workbook.save(file_name)


schedule.every(1).minutes.do(log_temperature_to_excel)


if __name__ == "__main__":
    print("Đang ghi nhận nhiệt độ CPU. Nhấn Ctrl+C để dừng.")
    while True:
        schedule.run_pending()
        time.sleep(1)

